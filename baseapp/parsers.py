import openpyxl as openpyxl
import pandas as pd

from .cruid import to_database, climat_parameters_to_database


def read_project_cost_info(sheet, row):
    dd = sheet.cell(row, 2).value
    info = None
    if dd:
        ddd = str(dd).strip().split('\n')
        if len(ddd) > 1:
            info = (ddd[0].strip(),
                    True if ddd[1].strip() == '(with project)' else False)
    return info


def read_col_scenario(sheet, start_row, end_row, col, need_round=False):
    list_value = []
    for row in range(start_row, end_row):
        value = sheet.cell(row, col).value
        if need_round:
            value = round(value, 2)
        list_value.append(value)
    return list_value


def read_scenario(sheet, scenario):
    projects = []
    costs = []
    for i in range(scenario['start_row'] + 3,
                   scenario['end_row'], 3):
        info = read_project_cost_info(sheet, i)
        if info:
            projects.append(info)
            if info[0] not in costs:
                costs.append(info[0])
    scenario['projects'] = projects
    scenario['costs'] = costs

    list_costs = []
    list_proj = []
    for cost, proj in scenario['projects']:
        for _ in range(3):
            list_costs.append(cost)
            list_proj.append(proj)

    start_row = scenario['start_row'] + 3
    end_row = scenario['end_row']
    df_dict = {
        'cost': pd.Series(list_costs),
        'with project': pd.Series(list_proj),
        'unit': pd.Series(read_col_scenario(sheet, start_row, end_row, 3))
    }
    for i in range(scenario['lifetime']):
        col = 4 + i
        df_dict[scenario['start_year'] + i] = pd.Series(
            read_col_scenario(sheet, start_row, end_row, col, need_round=True),
            dtype="float32")

    scenario['df'] = pd.DataFrame(df_dict)

    list_investment_costs = []
    row = scenario['start_row'] + 2
    for i in range(scenario['lifetime']):
        col = 4 + i
        year = scenario['start_year'] + i
        value = sheet.cell(row, col).value
        if value is None:
            value = 0
        value = round(value, 2)
        list_investment_costs.append({'year': year, 'value': value})
    scenario['investment_costs'] = list_investment_costs


def read_climat_parameters(project_data, wb):
    sheet = wb['Climate impacts']

    disaster_impacts = []
    level_disaster_impacts = []
    climate_impacts_type_value = []

    start_row = 35
    col = 3
    for row in range(start_row, start_row + 4):
        value = sheet.cell(row, col).value
        if value:
            disaster_impacts.append(str(value).strip())

    start_row = 42
    col = 2
    for row in range(start_row, start_row +24, 6):
        value = sheet.cell(row, col).value
        if value:
            level_disaster_impacts.append(str(value).strip())

    start_row = 44
    col = 2
    for row in range(start_row, start_row + 4):
        value = sheet.cell(row, col).value
        if value:
            climate_impacts_type_value.append({'name': str(value).strip(),
                                               'section': 'disaster'})

    start_row = 6
    col = 2
    for row in range(start_row, start_row + 28):
        value = sheet.cell(row, col).value
        if value:
            climate_impacts_type_value.append({'name': str(value).strip(),
                                               'section': 'climate_conditions'})

    project_data['disaster_impacts'] = disaster_impacts
    project_data['level_disaster_impacts'] = level_disaster_impacts
    project_data['climate_impacts_type_value'] = climate_impacts_type_value


def read_change_climate_condition(project_data, wb):

    sheet = wb['Climate impacts']
    start_row = 4
    start_col = 2
    year1 = int(sheet.cell(start_row, start_col+2).value)
    year2 = int(sheet.cell(start_row, start_col + 4).value)
    year_template = [(start_col + 2, year1), (start_col + 4, year2), ]
    impact_template = [(0, False), (1, True), ]
    with_project_data_template = [(0, True), (7, False), ]

    type_value = None
    type_value_with_cost = ['Expected change in quantity produced (%)',
                            'Expected change in price (%)']
    list_climate_condition = []
    count_cost = 0
    for row in range(start_row+2, start_row+28+2):
        cost = sheet.cell(row, start_col + 1).value
        current_type_value = sheet.cell(row, start_col).value
        if current_type_value:
            type_value = current_type_value
            count_cost = 0
        if type_value in type_value_with_cost:
            count_cost += 1
        else:
            cost = None
        if count_cost > 5:
            continue
        for col, year in year_template:
            for i_col, impact in impact_template:
                for p_col, with_project in with_project_data_template:
                    value = sheet.cell(row, col + i_col + p_col).value
                    list_climate_condition.append(
                        {'type_value': type_value,
                         'cost': cost,
                         'with_project': with_project,
                         'impact': impact,
                         'year': year,
                         'value': value})
    project_data['data_climate_condition'] = list_climate_condition
    climate_condition_dict = {}
    num = 0
    for x in list_climate_condition:
        # print(x)
        climate_condition_dict[num] = x
    df = pd.DataFrame(climate_condition_dict).T


def read_change_disaster_impact(project_data, wb):

    sheet = wb['Climate impacts']
    start_row = 39
    step_row = 28
    start_col = 2
    year_template = []
    for i in [2, 9, 11]:
        col = start_col+i
        year_template.append((int(sheet.cell(start_row+1, col).value), col))

    disaster_template = []
    for i in range(3):
        row = start_row + i * step_row
        s = str(sheet.cell(row, start_col).value).strip()
        disaster = s.split('(')[1].split(')')[0]
        disaster_template.append((disaster, row))

    level_template = []
    step_row = 6
    for i in range(4):
        row = start_row + i * step_row + 3
        level = str(sheet.cell(row, start_col).value).strip()
        return_period = int(sheet.cell(row+1, start_col+2).value)
        level_template.append((level, return_period, i * step_row + 4))

    impact_template = [(0, False), (1, True), ]

    list_change_disaster_impact = []
    for disaster_impact, disaster_row in disaster_template:
        for level, return_period, level_row in level_template:

            for i in range(5):
                row = disaster_row + level_row + i
                type_value = sheet.cell(row, start_col).value

                for year, year_col in year_template:
                    for i_col, impact in impact_template:
                        col = year_col + i_col
                        value = sheet.cell(row, col).value

                        list_change_disaster_impact.append(
                            {'disaster_impact': disaster_impact,
                             'level_disaster_impact': level,
                             'type_value': type_value,
                             'impact': impact,
                             'year': year,
                             'value': value})

    project_data['data_change_disaster_impact'] = list_change_disaster_impact


def read_from_exel(filename, project_name):
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb['Baseline scenario']

    optimistic_scenario = {'name': 'OPTIMISTIC BASELINE SCENARIO'}
    pesimistic_scenario = {'name': 'PESIMISTIC BASELINE SCENARIO'}

    work_scenario = optimistic_scenario
    for i in range(7, 200000):
        item_name = str(sheet.cell(i, 2).value).strip()
        if item_name == optimistic_scenario['name']:
            work_scenario = optimistic_scenario
            work_scenario['start_row'] = i
        elif item_name == pesimistic_scenario['name']:
            work_scenario = pesimistic_scenario
            work_scenario['start_row'] = i
        elif item_name == 'Discounted costs':
            work_scenario['end_row'] = i
            if work_scenario == pesimistic_scenario:
                break

    optimistic_scenario['end_row'] = optimistic_scenario['start_row'] + 33
    pesimistic_scenario['end_row'] = pesimistic_scenario['start_row'] + 33

    project_data = {'discount_rate': float(sheet.cell(5, 4).value),
                    'start_year': int(sheet.cell(4, 4).value),
                    'lifetime': 0,
                    'name': project_name,
                    'optimistic_scenario': optimistic_scenario,
                    'pesimistic_scenario': pesimistic_scenario}
    for i in range(100):
        dd = sheet.cell(8, 4+i).value
        if dd is None:
            project_data['lifetime'] = i
            break

    optimistic_scenario['lifetime'] = project_data['lifetime']
    pesimistic_scenario['lifetime'] = project_data['lifetime']
    optimistic_scenario['start_year'] = project_data['start_year']
    pesimistic_scenario['start_year'] = project_data['start_year']

    read_scenario(sheet, optimistic_scenario)
    read_scenario(sheet, pesimistic_scenario)

    read_climat_parameters(project_data, wb)

    read_change_climate_condition(project_data, wb)
    read_change_disaster_impact(project_data, wb)

    read_sensitivity_analysis(project_data, wb)

    to_database(project_data)


def read_sensitivity_analysis(project_data, wb):

    sheet = wb['Results']
    start_row = 81
    start_col = 2
    list_data = []
    for row in range(start_row, start_row+4):
        name = str(sheet.cell(row, start_col).value).strip()
        value = sheet.cell(row, start_col+6).value
        list_data.append(
            {'name': name,
             'value': value,
             'disaster_impact_name': '',
             'section': 'climate_conditions',
             'value': value})

    row = start_row + 4
    for disaster_impact_name in project_data['disaster_impacts']:
        for i in range(2):
            name = str(sheet.cell(row, start_col).value).strip().split('[')[0]
            value = sheet.cell(row, start_col + 6).value
            list_data.append(
                {'name': name,
                 'value': value,
                 'disaster_impact_name': disaster_impact_name,
                 'section': 'disaster',
                 'value': value})
            row += 1

    project_data['sensitivity_analysis'] = list_data
